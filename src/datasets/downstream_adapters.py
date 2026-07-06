from __future__ import annotations

import csv
import json
import math
import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from src.utils.sequence import iter_fasta_records, normalize_sequence


AFFINITY_PATTERN = re.compile(
    r"^\s*[A-Za-z]+\s*=\s*([-+]?[0-9]*\.?[0-9]+(?:[eE][-+]?[0-9]+)?)\s*([A-Za-zµμ]+)\s*$"
)
AFFINITY_UNIT_TO_MOLAR = {
    "pm": 1e-12,
    "nm": 1e-9,
    "um": 1e-6,
    "µm": 1e-6,
    "μm": 1e-6,
    "mm": 1e-3,
    "m": 1.0,
}


@dataclass(frozen=True)
class DownstreamSample:
    sample_id: str
    sequence: str
    label: str
    task_name: str
    split: str | None = None
    peptide_sequence: str | None = None
    pair_key: str | None = None


def read_structured_rows(path: str | Path) -> list[dict[str, Any]]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_delimited_rows(file_path, delimiter=",")
    if suffix == ".tsv":
        return _read_delimited_rows(file_path, delimiter="\t")
    if suffix == ".json":
        return _read_json_rows(file_path)
    if suffix == ".jsonl":
        return _read_jsonl_rows(file_path)
    if suffix == ".xlsx":
        return _read_xlsx_rows(file_path)

    raise ValueError(f"Unsupported structured file format: {file_path}")


def _read_delimited_rows(path: Path, *, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [_normalize_row(row) for row in reader]


def _read_json_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    if isinstance(payload, list):
        return [_coerce_json_row(item, path) for item in payload]
    if isinstance(payload, dict):
        return [_coerce_json_row(payload, path)]

    raise ValueError(f"JSON payload must be an object or array: {path}")


def _read_jsonl_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            payload = json.loads(stripped)
            rows.append(_coerce_json_row(payload, path, line_number))
    return rows


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading PPIKB .xlsx files requires openpyxl. Install dependencies with "
            "`pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = ["" if value is None else str(value).strip() for value in header_row]
    records: list[dict[str, Any]] = []
    for row in rows:
        record = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(_normalize_row(record))
    return records


def _coerce_json_row(payload: Any, path: Path, line_number: int | None = None) -> dict[str, Any]:
    if not isinstance(payload, dict):
        location = f"{path}:{line_number}" if line_number is not None else str(path)
        raise ValueError(f"JSON row must be an object at {location}")
    return _normalize_row(payload)


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized_key = str(key).strip()
        if isinstance(value, str):
            normalized[normalized_key] = value.strip()
        else:
            normalized[normalized_key] = value
    return normalized


def load_downstream_samples(
    task_name: str,
    input_path: str | Path,
    *,
    sample_limit: int | None = None,
) -> list[DownstreamSample]:
    normalized_task = task_name.strip().lower()
    if normalized_task == "toxteller":
        return load_toxteller_samples(input_path, sample_limit=sample_limit)
    if normalized_task == "conotoxin":
        return load_conotoxin_samples(input_path, sample_limit=sample_limit)
    if normalized_task == "prmftp":
        return load_prmftp_samples(input_path, sample_limit=sample_limit)
    if normalized_task == "ppikb":
        return load_ppikb_samples(input_path, sample_limit=sample_limit)

    raise ValueError(f"Unsupported task name: {task_name}")


def load_toxteller_samples(
    input_path: str | Path,
    *,
    sample_limit: int | None = None,
) -> list[DownstreamSample]:
    root = Path(input_path)
    fasta_files = _discover_toxteller_files(root)

    samples: list[DownstreamSample] = []
    for file_path in fasta_files:
        split = _parse_toxteller_split_from_filename(file_path.name)
        for index, record in enumerate(iter_fasta_records(file_path), start=1):
            label = _parse_toxteller_label_from_header(record.protein_id)
            samples.append(
                DownstreamSample(
                    sample_id=f"{file_path.stem}:{index}:{record.protein_id}",
                    sequence=record.sequence,
                    label=str(label),
                    task_name="toxteller",
                    split=split,
                )
            )
            if sample_limit is not None and len(samples) >= sample_limit:
                return samples

    return samples


def load_conotoxin_samples(
    input_path: str | Path,
    *,
    sample_limit: int | None = None,
) -> list[DownstreamSample]:
    root = Path(input_path)
    fasta_files = _discover_conotoxin_files(root)

    samples: list[DownstreamSample] = []
    for file_path in fasta_files:
        label = _parse_conotoxin_label_from_filename(file_path.name)
        for index, record in enumerate(iter_fasta_records(file_path), start=1):
            samples.append(
                DownstreamSample(
                    sample_id=f"{file_path.stem}:{index}:{record.protein_id}",
                    sequence=record.sequence,
                    label=str(label),
                    task_name="conotoxin",
                    split=None,
                )
            )
            if sample_limit is not None and len(samples) >= sample_limit:
                return samples

    return samples


def load_prmftp_samples(
    input_path: str | Path,
    *,
    sample_limit: int | None = None,
) -> list[DownstreamSample]:
    root = Path(input_path)
    fasta_files = _discover_prmftp_files(root)

    samples: list[DownstreamSample] = []
    for file_path in fasta_files:
        split = _parse_split_from_filename(file_path.name)
        records, skipped_records = _load_valid_fasta_records(file_path)
        for index, (protein_id, sequence) in enumerate(records, start=1):
            samples.append(
                DownstreamSample(
                    sample_id=f"{file_path.stem}:{index}",
                    sequence=sequence,
                    label=_parse_prmftp_label_bits(protein_id, file_path=file_path, record_index=index),
                    task_name="prmftp",
                    split=split,
                )
            )
            if sample_limit is not None and len(samples) >= sample_limit:
                return samples
        if skipped_records:
            warnings.warn(
                f"Skipped {skipped_records} malformed PrMFTP FASTA record(s) from {file_path}.",
                RuntimeWarning,
            )

    return samples


def load_ppikb_samples(
    input_path: str | Path,
    *,
    sample_limit: int | None = None,
) -> list[DownstreamSample]:
    root = Path(input_path)
    structured_files = _discover_ppikb_files(root)

    samples: list[DownstreamSample] = []
    for file_path in structured_files:
        relative_name = file_path.name if file_path == root else file_path.relative_to(root).as_posix()
        split = _parse_ppikb_split_from_filename(file_path.name)
        skipped_rows = 0
        for row_number, row in enumerate(read_structured_rows(file_path), start=1):
            row_id = str(row.get("ID") or f"row_{row_number}")
            sequence_value = row.get("Protein_Sequence")
            peptide_sequence_value = row.get("Peptide_Sequence")
            affinity_value = row.get("Affinity")

            if not isinstance(sequence_value, str) or not sequence_value:
                raise ValueError(f"Missing Protein_Sequence in {file_path}:{row_number}")
            if not isinstance(peptide_sequence_value, str) or not peptide_sequence_value:
                raise ValueError(f"Missing Peptide_Sequence in {file_path}:{row_number}")
            if not isinstance(affinity_value, str) or not affinity_value:
                raise ValueError(f"Missing Affinity in {file_path}:{row_number}")

            try:
                sample = _build_ppikb_sample(
                    relative_name=relative_name,
                    row_id=row_id,
                    sequence_value=sequence_value,
                    peptide_sequence_value=peptide_sequence_value,
                    affinity_value=affinity_value,
                    split=split,
                )
            except ValueError:
                skipped_rows += 1
                continue
            samples.append(sample)
            if sample_limit is not None and len(samples) >= sample_limit:
                return samples
            
        if skipped_rows:
            warnings.warn(
                f"Skipped {skipped_rows} unsupported or invalid PPIKB row(s) from {file_path}.",
                RuntimeWarning,
            )

    return samples


def _stable_sequence_hash(sequence: str) -> str:
    import hashlib
    return hashlib.sha256(str(sequence).encode("utf-8")).hexdigest()


def parse_affinity_to_molar_string(raw_affinity: str) -> str:
    match = AFFINITY_PATTERN.match(raw_affinity)
    if match is None:
        raise ValueError(f"Unsupported affinity format: {raw_affinity}")

    numeric_value = float(match.group(1))
    raw_unit = match.group(2).strip().lower()
    scale = AFFINITY_UNIT_TO_MOLAR.get(raw_unit)
    if scale is None:
        raise ValueError(f"Unsupported affinity unit: {raw_affinity}")

    return f"{numeric_value * scale:.12g}"


def parse_kd_to_pkd_string(raw_affinity: str) -> str:
    normalized = raw_affinity.replace("µ", "u").replace("μ", "u")
    lower_value = normalized.lower()

    if "pkd" in lower_value:
        match = re.search(r"(\d+(?:\.\d+)?)", normalized)
        if match is None:
            raise ValueError(f"Unsupported pKd format: {raw_affinity}")
        return f"{float(match.group(1)):.12g}"

    if "kd" not in lower_value:
        raise ValueError(f"PPIKB regression expects Kd affinity, got: {raw_affinity}")
    if any(token in lower_value for token in ("ki", "ic50", "ec50", "km", "ka", "pki", "pa2")):
        raise ValueError(f"PPIKB regression only accepts Kd affinity, got: {raw_affinity}")

    match = re.search(
        r"kd\s*=?\s*([-+]?[0-9]*\.?[0-9]+(?:e[-+]?[0-9]+)?)\s*([a-zA-Z]+)",
        normalized,
        re.IGNORECASE,
    )
    if match is None:
        raise ValueError(f"Unsupported Kd affinity format: {raw_affinity}")

    numeric_value = float(match.group(1))
    if numeric_value <= 0:
        raise ValueError(f"Kd affinity must be positive: {raw_affinity}")
    raw_unit = match.group(2).strip().lower()
    scale = AFFINITY_UNIT_TO_MOLAR.get(raw_unit)
    if scale is None:
        raise ValueError(f"Unsupported Kd affinity unit: {raw_affinity}")

    molar_value = numeric_value * scale
    if molar_value <= 0:
        raise ValueError(f"Kd affinity must be positive: {raw_affinity}")
    return f"{-math.log10(molar_value):.12g}"


def _discover_toxteller_files(root: Path) -> list[Path]:
    if root.is_file():
        return [root]

    preferred_files = [
        root / "training_dataset.fasta",
        root / "independent_dataset.fasta",
    ]
    existing_preferred = [path for path in preferred_files if path.exists()]
    if existing_preferred:
        return existing_preferred

    files = sorted((path for path in root.glob("*.fasta")), key=lambda path: path.as_posix())
    if not files:
        raise FileNotFoundError(f"No ToxTeller FASTA files found in {root}")
    return files


def _discover_conotoxin_files(root: Path) -> list[Path]:
    if root.is_file():
        return [root]

    preferred_files = [root / "pos.fasta", root / "neg.fasta"]
    existing_preferred = [path for path in preferred_files if path.exists()]
    if existing_preferred:
        if len(existing_preferred) != len(preferred_files):
            missing = [path.name for path in preferred_files if not path.exists()]
            raise FileNotFoundError(f"Incomplete conotoxin FASTA set under {root}; missing {missing}.")
        return preferred_files

    files = sorted((path for path in root.glob("*.fasta")), key=lambda path: path.as_posix())
    if not files:
        raise FileNotFoundError(f"No conotoxin FASTA files found in {root}")
    return files


def _discover_prmftp_files(root: Path) -> list[Path]:
    if root.is_file():
        return [root]

    preferred_files = [root / "train.txt", root / "test.txt"]
    existing_preferred = [path for path in preferred_files if path.exists()]
    if existing_preferred:
        return existing_preferred

    files = sorted((path for path in root.glob("*.txt")), key=lambda path: path.as_posix())
    if not files:
        raise FileNotFoundError(f"No PrMFTP FASTA-like files found in {root}")
    return files


def _discover_ppikb_files(root: Path) -> list[Path]:
    if root.is_file():
        return [root]

    if root.name == "regression":
        default_run_dir = root / "run_1"
        if default_run_dir.exists():
            return _discover_ppikb_regression_split_files(default_run_dir)

    regression_dir = root / "regression"
    default_run_dir = regression_dir / "run_1"
    if default_run_dir.exists():
        return _discover_ppikb_regression_split_files(default_run_dir)

    if root.name.startswith("run_") and root.is_dir():
        split_files = _discover_ppikb_regression_split_files(root)
        if split_files:
            return split_files

    canonical_csv = root / "PPIKB.csv"
    if canonical_csv.exists():
        return [canonical_csv]

    search_root = regression_dir if regression_dir.exists() else root
    files = sorted(
        (
            path
            for path in search_root.rglob("*")
            if path.is_file() and path.suffix.lower() in {".csv", ".tsv", ".json", ".jsonl", ".xlsx"}
            and ".ipynb_checkpoints" not in path.parts
            and "preprocessed" not in path.parts
        ),
        key=lambda path: path.as_posix(),
    )
    if not files:
        raise FileNotFoundError(f"No supported PPIKB files found in {root}")
    return files


def _discover_ppikb_regression_split_files(run_dir: Path) -> list[Path]:
    split_files = [
        run_dir / "train_set.csv",
        run_dir / "val_set.csv",
        run_dir / "test_set.csv",
    ]
    existing_split_files = [path for path in split_files if path.exists()]
    if existing_split_files and len(existing_split_files) != len(split_files):
        missing = [path.name for path in split_files if not path.exists()]
        raise FileNotFoundError(
            f"Incomplete PPIKB regression split under {run_dir}; missing {missing}."
        )
    return existing_split_files


def _load_valid_fasta_records(path: Path) -> tuple[list[tuple[str, str]], int]:
    try:
        records = [(record.protein_id, record.sequence) for record in iter_fasta_records(path)]
        return records, 0
    except ValueError:
        # Fall back to record-by-record parsing so one malformed FASTA entry
        # does not abort the whole downstream manifest build.
        current_header: str | None = None
        current_sequence_lines: list[str] = []
        valid_records: list[tuple[str, str]] = []
        skipped_records = 0

        def flush_record() -> None:
            nonlocal current_header, current_sequence_lines, skipped_records
            if current_header is None:
                return
            protein_id = current_header.split(None, 1)[0]
            raw_sequence = "".join(current_sequence_lines)
            current_header = None
            current_sequence_lines = []
            try:
                sequence = normalize_sequence(raw_sequence)
            except ValueError:
                skipped_records += 1
                return
            valid_records.append((protein_id, sequence))

        with path.open("r", encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, start=1):
                stripped = line.strip()
                if not stripped:
                    continue
                if stripped.startswith(">"):
                    flush_record()
                    current_header = stripped[1:].strip()
                    current_sequence_lines = []
                    continue
                if current_header is None:
                    raise ValueError(
                        f"FASTA sequence line appears before a header at {path}:{line_number}"
                    )
                current_sequence_lines.append(stripped)

        flush_record()
        return valid_records, skipped_records


def _build_ppikb_sample(
    *,
    relative_name: str,
    row_id: str,
    sequence_value: str,
    peptide_sequence_value: str,
    affinity_value: str,
    split: str | None,
) -> DownstreamSample:
    protein_sequence = normalize_sequence(sequence_value)
    peptide_sequence = normalize_sequence(peptide_sequence_value)
    protein_hash = _stable_sequence_hash(protein_sequence)
    peptide_hash = _stable_sequence_hash(peptide_sequence)
    return DownstreamSample(
        sample_id=f"{relative_name}:{row_id}",
        sequence=protein_sequence,
        label=parse_kd_to_pkd_string(affinity_value),
        task_name="ppikb",
        split=split,
        peptide_sequence=peptide_sequence,
        pair_key=f"{protein_hash}::{peptide_hash}",
    )


def _row_alias_value(row: dict[str, Any], aliases: tuple[str, ...]) -> Any | None:
    lowercase_row = {str(key).strip().lower(): value for key, value in row.items()}
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
        lowered_alias = alias.strip().lower()
        if lowered_alias in lowercase_row and lowercase_row[lowered_alias] not in (None, ""):
            return lowercase_row[lowered_alias]
    return None


def _required_row_alias_value(
    row: dict[str, Any],
    aliases: tuple[str, ...],
    *,
    file_path: Path,
    row_number: int,
    field_name: str,
) -> Any:
    value = _row_alias_value(row, aliases)
    if value is None or not str(value).strip():
        raise ValueError(
            f"Missing {field_name} in {file_path}:{row_number}. "
            f"Expected one of columns {list(aliases)}."
        )
    return value


def _parse_toxteller_label_from_header(header_value: str) -> int:
    normalized = str(header_value).strip().lower()
    if normalized.startswith("pos"):
        return 1
    if normalized.startswith("neg"):
        return 0
    raise ValueError(
        "ToxTeller label could not be inferred from FASTA header. "
        f"Expected a 'pos' or 'neg' prefix in {header_value!r}."
    )


def _parse_conotoxin_label_from_filename(filename: str) -> int:
    stem = Path(filename).stem.lower()
    if stem in {"pos", "positive"} or stem.startswith("pos_") or stem.startswith("positive_"):
        return 1
    if stem in {"neg", "negative"} or stem.startswith("neg_") or stem.startswith("negative_"):
        return 0
    raise ValueError(
        "Conotoxin label could not be inferred from filename. "
        f"Expected a pos/positive or neg/negative FASTA filename, got {filename!r}."
    )


def _parse_split_from_filename(filename: str) -> str | None:
    lower_name = filename.lower()
    if "train" in lower_name:
        return "train"
    if "test" in lower_name:
        return "test"
    if "val" in lower_name or "valid" in lower_name:
        return "val"
    return None


def _parse_toxteller_split_from_filename(filename: str) -> str | None:
    lower_name = filename.lower()
    if "training_dataset" in lower_name:
        return "train"
    if "independent_dataset" in lower_name:
        return "test"
    return _parse_split_from_filename(filename)


def _parse_ppikb_split_from_filename(filename: str) -> str | None:
    lower_name = filename.lower()
    if lower_name == "train_set.csv" or lower_name.startswith("train_"):
        return "train"
    if lower_name == "val_set.csv" or lower_name.startswith("val_") or lower_name.startswith("valid_"):
        return "val"
    if lower_name == "test_set.csv" or lower_name.startswith("test_"):
        return "test"
    return _parse_split_from_filename(filename)


def _parse_prmftp_label_bits(header_value: str, *, file_path: Path, record_index: int) -> str:
    label_bits = str(header_value).strip().split(None, 1)[0]
    if len(label_bits) != 21 or any(bit not in {"0", "1"} for bit in label_bits):
        raise ValueError(
            f"Expected a 21-bit PrMFTP multilabel header in {file_path}:{record_index}, "
            f"got {header_value!r}."
        )
    return label_bits
